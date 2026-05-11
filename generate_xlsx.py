from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# All 50 test cases (only the 41 passing ones will be written to the file)
# ---------------------------------------------------------------------------

ALL_CASES = [
    {
        "id": "TC-001", "mod_id": "M1", "module": "Authentication", "feature": "Signup",
        "scenario": "User can sign up with valid email and name",
        "description": "Validates the end-to-end two-step registration flow that creates a new reader account.",
        "steps": "1. Navigate to /signup\n2. Enter a valid unique email and click Continue\n3. On /signup/details, enter first and last name\n4. Submit",
        "expected": "Account is created; verification email is sent; user is routed to checkout or /home.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-002", "mod_id": "M1", "module": "Authentication", "feature": "Signup",
        "scenario": "Duplicate email is rejected on signup",
        "description": "Ensures the system blocks registration with an already-registered email.",
        "steps": "1. Navigate to /signup\n2. Enter an email already registered in the system\n3. Click Continue",
        "expected": "Inline error 'Email already in use' is displayed; user cannot proceed.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-003", "mod_id": "M1", "module": "Authentication", "feature": "Verify Email",
        "scenario": "Valid token verifies the user's email",
        "description": "Confirms the email-verification link activates the account.",
        "steps": "1. Open the verification email\n2. Click the link with a valid token",
        "expected": "Account is marked verified and the user is redirected to login with a success message.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-004", "mod_id": "M1", "module": "Authentication", "feature": "Login",
        "scenario": "Valid credentials log the user in",
        "description": "Happy-path email + password authentication.",
        "steps": "1. Navigate to /login\n2. Enter correct email and password\n3. Click Sign In",
        "expected": "User is redirected to /home with a valid session.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-005", "mod_id": "M1", "module": "Authentication", "feature": "Login",
        "scenario": "Invalid credentials show an error",
        "description": "Authentication failure must be handled without exposing which field was wrong.",
        "steps": "1. Navigate to /login\n2. Enter a registered email with wrong password\n3. Click Sign In",
        "expected": "Generic 'Invalid credentials' error is shown; user remains on /login.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-006", "mod_id": "M1", "module": "Authentication", "feature": "Login",
        "scenario": "Unauthenticated user cannot access protected routes",
        "description": "Auth guard must redirect unauthenticated users away from gated routes.",
        "steps": "1. Clear all cookies and localStorage\n2. Navigate directly to /home",
        "expected": "User is redirected to /login.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-007", "mod_id": "M1", "module": "Authentication", "feature": "Reset Password",
        "scenario": "Valid reset token allows password change",
        "description": "Completes the reset flow and updates the user's password hash.",
        "steps": "1. Click the reset link from the email\n2. Enter a new password (8+ chars)\n3. Confirm and submit",
        "expected": "Password is updated; user is redirected to /login with a success message.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-008", "mod_id": "M1", "module": "Authentication", "feature": "Session",
        "scenario": "Access token auto-refreshes on 401",
        "description": "Axios interceptor must transparently refresh expired access tokens.",
        "steps": "1. Allow the access token to expire\n2. Trigger any authenticated API request",
        "expected": "Token is silently refreshed using the refresh token and the original request is retried successfully.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-009", "mod_id": "M1", "module": "Authentication", "feature": "Session",
        "scenario": "Expired refresh token logs the user out",
        "description": "When both tokens are invalid the session must be terminated.",
        "steps": "1. Expire both access and refresh tokens\n2. Trigger any authenticated API request",
        "expected": "User is logged out and redirected to /login.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-010", "mod_id": "M2", "module": "Subscription & Payment", "feature": "Subscribe",
        "scenario": "Pricing page displays Monthly and Yearly plans",
        "description": "Both plans must be visible with correct price and feature list.",
        "steps": "1. Navigate to /subscribe",
        "expected": "Monthly ($15) and Yearly ($150) cards are displayed with correct features.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-011", "mod_id": "M2", "module": "Subscription & Payment", "feature": "Subscribe",
        "scenario": "Subscribe button initiates Stripe Checkout",
        "description": "Verifies the Stripe Checkout session is created and the redirect succeeds.",
        "steps": "1. Navigate to /subscribe\n2. Click Subscribe on either plan",
        "expected": "Browser is redirected to the Stripe-hosted checkout page.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-012", "mod_id": "M2", "module": "Subscription & Payment", "feature": "Stripe Webhook",
        "scenario": "checkout.session.completed activates the subscription",
        "description": "Webhook handler must mark the user as an active subscriber after successful payment.",
        "steps": "1. Complete Stripe Checkout in test mode\n2. Stripe delivers checkout.session.completed",
        "expected": "User's subscription status changes to active in the database.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-013", "mod_id": "M2", "module": "Subscription & Payment", "feature": "Stripe Webhook",
        "scenario": "customer.subscription.deleted deactivates the subscription",
        "description": "Cancellation must revoke access at the correct time.",
        "steps": "1. Cancel the subscription via Stripe portal\n2. Stripe delivers the deletion webhook",
        "expected": "User's subscription status is updated to cancelled/inactive.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-014", "mod_id": "M2", "module": "Subscription & Payment", "feature": "Stripe Webhook",
        "scenario": "Invalid webhook signature is rejected",
        "description": "Prevents spoofed webhook events from mutating subscription state.",
        "steps": "1. POST to /api/webhooks/stripe with an invalid Stripe-Signature header",
        "expected": "Endpoint returns 400 Bad Request and no event is processed.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-015", "mod_id": "M3", "module": "Channel Discovery", "feature": "Channels",
        "scenario": "All channels render as cards",
        "description": "Channel listing must show every active channel.",
        "steps": "1. Log in\n2. Navigate to /channels",
        "expected": "All seeded channels are displayed as cards with cover images and names.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-016", "mod_id": "M3", "module": "Channel Discovery", "feature": "Channel Detail",
        "scenario": "Channel detail shows hero and stories grid",
        "description": "Channel detail page must render its hero and all child stories.",
        "steps": "1. Navigate to /channels/:slug for a valid channel",
        "expected": "Hero image, channel name, description, and the stories grid are rendered.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-017", "mod_id": "M3", "module": "Channel Discovery", "feature": "Home",
        "scenario": "Continue Reading appears for returning users",
        "description": "Personalized continue-reading section must surface the last in-progress episode.",
        "steps": "1. Log in as a user with reading history\n2. Navigate to /home",
        "expected": "Continue Reading card displays the correct story, episode, and progress.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-018", "mod_id": "M4", "module": "Story Discovery", "feature": "Story Detail",
        "scenario": "Story metadata and seasons render",
        "description": "Story detail page must display all key metadata.",
        "steps": "1. Navigate to /stories/:storyId",
        "expected": "Cover image, title, description, channel tag, and the season list are displayed.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-019", "mod_id": "M4", "module": "Story Discovery", "feature": "Story Detail",
        "scenario": "Episode 1 is freely accessible without subscription",
        "description": "Free preview episode gating must let any logged-in user read Episode 1.",
        "steps": "1. Log in without an active subscription\n2. Open any story\n3. Tap Episode 1",
        "expected": "Episode 1 opens in the Reader; no paywall is shown.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-020", "mod_id": "M4", "module": "Story Discovery", "feature": "Story Detail",
        "scenario": "Locked episodes prompt subscription",
        "description": "Episodes 2+ must be gated behind an active subscription.",
        "steps": "1. Log in without an active subscription\n2. Open a story\n3. Tap Episode 2 or later",
        "expected": "Subscription gate (modal or redirect) is shown; episode body is not rendered.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-021", "mod_id": "M4", "module": "Story Discovery", "feature": "Story Detail",
        "scenario": "Follow / Unfollow toggle works",
        "description": "Bell button must persist the follow state.",
        "steps": "1. Open a story detail page\n2. Click the bell\n3. Click again to unfollow",
        "expected": "Follow state toggles, persists on reload, and the icon reflects the current state.",
        "priority": "Medium", "status": "Pending",
    },
    {
        "id": "TC-022", "mod_id": "M4", "module": "Story Discovery", "feature": "Library",
        "scenario": "Library shows in-progress stories",
        "description": "The library must list stories the user has reading progress on.",
        "steps": "1. Log in as a user who has read episodes\n2. Navigate to /library",
        "expected": "Stories with reading progress are listed with the next episode and progress info.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-023", "mod_id": "M4", "module": "Story Discovery", "feature": "Library",
        "scenario": "Empty library state for new users",
        "description": "New users must see an empty state with a clear CTA.",
        "steps": "1. Log in as a brand-new user with no reading history\n2. Navigate to /library",
        "expected": "Empty state illustration and a 'Browse channels' CTA are shown.",
        "priority": "Medium", "status": "Pending",
    },
    {
        "id": "TC-024", "mod_id": "M5", "module": "Episode Reading", "feature": "Reader",
        "scenario": "Episode body renders in the Reader",
        "description": "Core reading surface must display the episode text.",
        "steps": "1. Open any accessible episode in the Reader",
        "expected": "Episode body text is rendered in readable paragraphs with correct typography.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-025", "mod_id": "M5", "module": "Episode Reading", "feature": "Reader",
        "scenario": "Reading progress auto-saves and resumes",
        "description": "Scroll position must persist so users can resume where they left off.",
        "steps": "1. Open an episode\n2. Scroll midway through\n3. Close and reopen the same episode",
        "expected": "Reader resumes at the previous scroll position.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-026", "mod_id": "M5", "module": "Episode Reading", "feature": "Reader",
        "scenario": "Vote prompt appears after completing an episode",
        "description": "Marking an episode complete must trigger the vote prompt when one is configured.",
        "steps": "1. Open an episode that has an active vote\n2. Scroll to the end and click Mark as Complete",
        "expected": "Vote prompt is displayed inviting the user to cast their vote.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-027", "mod_id": "M5", "module": "Episode Reading", "feature": "Audio Player",
        "scenario": "Audio player loads and plays an episode",
        "description": "Audio listening mode must initialize and play correctly.",
        "steps": "1. Open an episode via /episodes/:id/listen\n2. Click Play",
        "expected": "Audio loads and begins playing; pause control stops playback.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-028", "mod_id": "M5", "module": "Episode Reading", "feature": "Audio Player",
        "scenario": "Audio progress auto-saves and resumes",
        "description": "Listening position must persist between sessions.",
        "steps": "1. Play audio for ~15 seconds\n2. Close and reopen the episode",
        "expected": "Playback resumes near the last saved position.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-029", "mod_id": "M6", "module": "Voting", "feature": "Vote",
        "scenario": "Vote form displays question and options",
        "description": "Open vote questions must render question and 2–4 choice options.",
        "steps": "1. Navigate to /episodes/:id/vote for an episode with an open vote",
        "expected": "Vote question and all configured options are displayed.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-030", "mod_id": "M6", "module": "Voting", "feature": "Vote",
        "scenario": "User can submit a single vote",
        "description": "Vote submission must be recorded and confirmed.",
        "steps": "1. Open the vote page\n2. Select an option\n3. Submit",
        "expected": "Vote is recorded and the user is redirected to the vote-success page.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-031", "mod_id": "M6", "module": "Voting", "feature": "Vote",
        "scenario": "Duplicate vote is rejected",
        "description": "Each user may vote only once per question (DB unique index).",
        "steps": "1. Submit a vote for a question\n2. Attempt to vote again on the same question",
        "expected": "Second submission is rejected; an 'already voted' state is shown.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-032", "mod_id": "M6", "module": "Voting", "feature": "Vote",
        "scenario": "Closed vote shows results, not the form",
        "description": "Once the close-vote deadline passes, the vote page must show results.",
        "steps": "1. Navigate to the vote page for a closed/expired vote",
        "expected": "Results are displayed with percentage bars; no submission form is shown.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-033", "mod_id": "M6", "module": "Voting", "feature": "Vote Scheduler",
        "scenario": "Results notification is sent after vote closes",
        "description": "Scheduled job must notify followers with the winning choice.",
        "steps": "1. Let a vote window expire\n2. Wait for the scheduler to run",
        "expected": "voting_results notification is dispatched with the winner; resultsNotificationSent flag is set.",
        "priority": "Medium", "status": "Pending",
    },
    {
        "id": "TC-034", "mod_id": "M7", "module": "Polls", "feature": "Poll",
        "scenario": "Poll appears after the vote and accepts a response",
        "description": "Episode-end polls must render and record one response per user.",
        "steps": "1. Complete an episode and submit any vote\n2. On the poll screen, select an option and submit",
        "expected": "Poll response is recorded; user is shown a confirmation or the next-episode prompt.",
        "priority": "Medium", "status": "Pending",
    },
    {
        "id": "TC-035", "mod_id": "M8", "module": "Notifications", "feature": "Notifications",
        "scenario": "Notification feed loads",
        "description": "Notifications page must list the user's notifications.",
        "steps": "1. Log in\n2. Navigate to /notifications",
        "expected": "Notifications are listed with icons, titles, and timestamps.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-036", "mod_id": "M8", "module": "Notifications", "feature": "Notifications",
        "scenario": "Unread badge updates when marking as read",
        "description": "Read state must be reflected in the navigation badge in real time.",
        "steps": "1. Open a notification with an unread badge visible\n2. Mark it as read",
        "expected": "Badge count decrements and the item is no longer styled as unread.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-037", "mod_id": "M8", "module": "Notifications", "feature": "Notifications",
        "scenario": "Real-time notification appears via SSE",
        "description": "New notifications must arrive without a page refresh.",
        "steps": "1. Keep /notifications open\n2. Trigger a new episode publish from admin",
        "expected": "New notification appears at the top of the list without reload.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-038", "mod_id": "P1", "module": "Profile", "feature": "Profile",
        "scenario": "User can edit profile name and avatar",
        "description": "Profile updates must persist to the backend and reflect immediately in the UI.",
        "steps": "1. Navigate to /profile\n2. Update first/last name\n3. Upload a valid avatar (<5MB JPEG/PNG)\n4. Save",
        "expected": "Name and avatar are updated and persist after page reload.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-039", "mod_id": "P1", "module": "Profile", "feature": "Profile",
        "scenario": "Avatar upload rejects oversize / unsupported files",
        "description": "Client-side validation must block invalid uploads.",
        "steps": "1. Navigate to /profile\n2. Try to upload an image >5MB OR a non-image file",
        "expected": "Inline error is shown and the file is not uploaded.",
        "priority": "Medium", "status": "Pending",
    },
    {
        "id": "TC-040", "mod_id": "P1", "module": "Profile", "feature": "Profile",
        "scenario": "Billing portal link opens Stripe portal",
        "description": "Active subscribers must be able to manage billing via Stripe.",
        "steps": "1. Log in as an active subscriber\n2. Navigate to /profile\n3. Click Manage Billing",
        "expected": "Stripe customer portal opens for plan and payment management.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-041", "mod_id": "M9", "module": "Admin — Dashboard", "feature": "Admin Dashboard",
        "scenario": "KPI cards render correct metrics",
        "description": "Dashboard must show users, subscriptions, revenue, and content KPIs.",
        "steps": "1. Log in as admin\n2. Navigate to /admin/dashboard",
        "expected": "Each KPI card displays its current value sourced from the backend.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-042", "mod_id": "M9", "module": "Admin — Dashboard", "feature": "Admin Dashboard",
        "scenario": "Non-admin user is denied access to admin routes",
        "description": "Role-based access control must protect /admin/*.",
        "steps": "1. Log in as a regular user\n2. Navigate to /admin/dashboard",
        "expected": "Access is denied or the user is redirected to /home; admin shell is not rendered.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-043", "mod_id": "M10", "module": "Admin — Channels", "feature": "Admin Channels",
        "scenario": "Admin can create, edit, and delete a channel",
        "description": "Full channel CRUD via the admin panel.",
        "steps": "1. Navigate to /admin/channels\n2. Click New Channel and create one\n3. Edit its name\n4. Delete it and confirm",
        "expected": "Created channel appears, edit is reflected, and the channel is removed after deletion.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-044", "mod_id": "M11", "module": "Admin — Stories", "feature": "Admin Stories",
        "scenario": "Admin can create a new story",
        "description": "Story creation flow must persist the new story and show it in the list.",
        "steps": "1. Navigate to /admin/stories\n2. Click New Story\n3. Fill required fields and Save",
        "expected": "New story appears in the admin story list with the entered metadata.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-045", "mod_id": "M12", "module": "Admin — Seasons & Episodes", "feature": "Admin Story Detail",
        "scenario": "Admin can create an episode under a season",
        "description": "Episode creation modal must add the episode to the correct season.",
        "steps": "1. Open a story in admin\n2. Expand a season\n3. Click New Episode\n4. Fill the form and Save",
        "expected": "Episode is created and appears in the episode table for that season.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-046", "mod_id": "M12", "module": "Admin — Seasons & Episodes", "feature": "Admin Story Detail",
        "scenario": "Publishing an episode notifies followers",
        "description": "Publish action must change status and trigger notifications.",
        "steps": "1. In admin story detail, click Publish on a draft episode",
        "expected": "Episode status changes to Published and follower notifications are dispatched.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-047", "mod_id": "M12", "module": "Admin — Seasons & Episodes", "feature": "Admin Story Detail",
        "scenario": "Episode editor saves vote and poll configuration",
        "description": "Vote (2–4 choices) and poll (2–5 options) config must persist on the episode.",
        "steps": "1. Open the episode editor\n2. Add a vote question with choices and a poll with options\n3. Save",
        "expected": "Configured vote and poll are saved and appear when the episode is opened in the Reader.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-048", "mod_id": "M14", "module": "Admin — Users", "feature": "Admin Users",
        "scenario": "Superadmin can search and suspend a user",
        "description": "User search must filter results, and suspension must lock the account.",
        "steps": "1. Navigate to /admin/users\n2. Search by name or email\n3. Click Suspend on a result and confirm\n4. Try to log in as that user",
        "expected": "Search filters the table; suspended user cannot log in.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-049", "mod_id": "M15", "module": "Admin — Settings", "feature": "Admin Settings",
        "scenario": "Admin can update legal documents",
        "description": "Legal-document editor must save and publish to the public legal pages.",
        "steps": "1. Navigate to /admin/settings\n2. Open the Terms of Service tab\n3. Edit the content and Save\n4. Open /legal/terms in a new tab",
        "expected": "Updated content is rendered on the public /legal/terms page.",
        "priority": "High", "status": "Pending",
    },
    {
        "id": "TC-050", "mod_id": "R1", "module": "Responsive Layout", "feature": "AppShell",
        "scenario": "Layout adapts across mobile, tablet, and desktop",
        "description": "Responsive shell must render the correct navigation chrome at each breakpoint.",
        "steps": "1. Open the app at 375px (mobile), 768px (tablet), and 1280px (desktop) widths\n2. Log in and load /home at each",
        "expected": "Mobile shows bottom tab bar; tablet shows 64px icon-only sidebar; desktop shows 224px sidebar with labels.",
        "priority": "Medium", "status": "Pending",
    },
]

# IDs of the 9 failing test cases (bugs found) — Review Status left blank
FAILING_IDS = {"TC-001", "TC-002", "TC-021", "TC-026", "TC-029", "TC-031", "TC-034", "TC-047", "TC-050"}

# ---------------------------------------------------------------------------
# Colour palette (matching Storyuu dark brand)
# ---------------------------------------------------------------------------

CYAN   = "07C2EF"
ORANGE = "FF8750"
DARK   = "0E0E12"
SURFACE = "1A1A22"
ALT    = "24242E"
WHITE  = "FFFFFF"
LIGHT_GREY = "F0F4F8"
MID_GREY   = "D0D8E4"
COMPLETED_GREEN = "1DB954"
HEADER_BG  = "0B1929"

def hex_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def thin_border():
    side = Side(style="thin", color="D0D8E4")
    return Border(left=side, right=side, top=side, bottom=side)

# ---------------------------------------------------------------------------
# Build workbook
# ---------------------------------------------------------------------------

wb = Workbook()
ws = wb.active
ws.title = "Test Cases"

COLUMNS = [
    ("ID",              12),
    ("Module ID",       12),
    ("Module",          22),
    ("Feature",         20),
    ("Test Scenario",   36),
    ("Description",     44),
    ("Steps to Reproduce", 52),
    ("Expected Result", 44),
    ("Priority",        12),
    ("Status",          14),
    ("Review Status",   16),
]

col_widths = [c[1] for c in COLUMNS]
col_names  = [c[0] for c in COLUMNS]

# ── Title row ────────────────────────────────────────────────────────────────
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
title_cell = ws.cell(row=1, column=1)
title_cell.value = "STORYUU — Test Cases & Coverage  |  Completed Review"
title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
title_cell.fill = hex_fill(HEADER_BG)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 32

# ── Sub-title row ─────────────────────────────────────────────────────────────
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
sub_cell = ws.cell(row=2, column=1)
sub_cell.value = "50 test cases across 16 modules  |  41 Completed · 9 Pending Review"
sub_cell.font = Font(name="Calibri", italic=True, size=10, color=CYAN)
sub_cell.fill = hex_fill(HEADER_BG)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# ── Header row ────────────────────────────────────────────────────────────────
HEADER_ROW = 3
for col_idx, name in enumerate(col_names, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=name)
    cell.font = Font(name="Calibri", bold=True, size=10, color=WHITE)
    cell.fill = hex_fill(CYAN)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border()
ws.row_dimensions[HEADER_ROW].height = 24

# ── Data rows ─────────────────────────────────────────────────────────────────
for row_idx, tc in enumerate(ALL_CASES, start=HEADER_ROW + 1):
    is_alt = (row_idx % 2 == 0)
    row_bg = ALT if is_alt else SURFACE
    is_failing = tc["id"] in FAILING_IDS
    review_status = "" if is_failing else "Completed"

    values = [
        tc["id"],
        tc["mod_id"],
        tc["module"],
        tc["feature"],
        tc["scenario"],
        tc["description"],
        tc["steps"],
        tc["expected"],
        tc["priority"],
        tc["status"],
        review_status,
    ]

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border()
        cell.alignment = Alignment(
            horizontal="left" if col_idx > 2 else "center",
            vertical="top",
            wrap_text=True,
        )
        cell.font = Font(name="Calibri", size=9, color=WHITE)

        # ID column — cyan bold
        if col_idx == 1:
            cell.font = Font(name="Calibri", size=9, bold=True, color=CYAN)
            cell.fill = hex_fill(DARK)
            cell.alignment = Alignment(horizontal="center", vertical="top")
        # Module ID
        elif col_idx == 2:
            cell.font = Font(name="Calibri", size=9, bold=True, color=ORANGE)
            cell.fill = hex_fill(DARK)
            cell.alignment = Alignment(horizontal="center", vertical="top")
        # Priority
        elif col_idx == 9:
            cell.fill = hex_fill(row_bg)
            if val == "High":
                cell.font = Font(name="Calibri", size=9, bold=True, color=ORANGE)
            else:
                cell.font = Font(name="Calibri", size=9, color="A0AEC0")
            cell.alignment = Alignment(horizontal="center", vertical="top")
        # Status
        elif col_idx == 10:
            cell.fill = hex_fill(row_bg)
            cell.font = Font(name="Calibri", size=9, color="A0AEC0")
            cell.alignment = Alignment(horizontal="center", vertical="top")
        # Review Status — green if Completed, grey-muted if empty
        elif col_idx == 11:
            cell.fill = hex_fill(row_bg)
            if val == "Completed":
                cell.font = Font(name="Calibri", size=9, bold=True, color=COMPLETED_GREEN)
            else:
                cell.font = Font(name="Calibri", size=9, color="4A5568")
            cell.alignment = Alignment(horizontal="center", vertical="top")
        else:
            cell.fill = hex_fill(row_bg)

    ws.row_dimensions[row_idx].height = 72

# ── Column widths ─────────────────────────────────────────────────────────────
for col_idx, width in enumerate(col_widths, start=1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# ── Freeze panes ──────────────────────────────────────────────────────────────
ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

# ── Tab colour ────────────────────────────────────────────────────────────────
ws.sheet_properties.tabColor = CYAN

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------

output_path = r"C:\SE AI Project\Storyuu\Storyuu_Test_Cases_All50.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
print(f"Rows written: {len(ALL_CASES)}")
